#!/usr/bin/python
# -*- coding: UTF-8 -*-
import sqlite3
from elasticsearch import Elasticsearch
from elasticsearch import helpers
from elasticsearch.exceptions import RequestError
import xlrd,openpyxl,pprint,os,datetime,re
from openpyxl.styles import PatternFill
from cfg.config import *



# GET /*
#
# DELETE /*
#
# GET _cat/indices?v
#
# GET _search
# {
#   "query": {
#     "match_all": {}
#   }
# }
#
# GET _search
# {
#   "query":{
#           "query_string":{
#               "default_field":"vus",
#               "query":" This function will provide suggestions"
#           }
#       },
#   "from":0,
#   "size":10
# }
#
# GET _search
# {
#   "query":{
#           "match_phrase":{
#               "vus":"Weather reminder"
#           }
#       }
# }
#
# GET /python_es01/doc/_search
# {
#   "suggest": {
#     "my_suggest": {
#       "text": "Weather remider",
#       "phrase": {
#         "field": "vus"
#       }
#     }
#   }
# }
#
# GET /python_es01/doc/_search
# {
#   "explain": true,
#   "query": {
#     "match": {
#       "vus": {
#         "query": "favorite",
#         "fuzziness": "AUTO"
#       }
#     }
#   },
#   "size":2
# }
#
# GET /python_es01/doc/_search
# {
#   "query": {
#     "term": {
#       "exact_value": {
#           "vus":"weather"
#       }
#     }
#   }
# }
#
# GET /python_es01/doc/_search
# { "query": { "bool": { "must": { "term": { "vus": "Weather reminder" } } } } }
#
# GET python_es01/_settings
#
# PUT /python_es01/doc/4
# {
#   "id" : 21322,
#   "app" : "HiBoard",
#   "string_id" : "collect_toast",
#   "vus" : "Weather Weather",
#   "vcn" : "看",
#   "vtr" : "F"
# }


class ElasticObj:
    def __init__(self, ip, index_name, index_type):
        # index=database/type=table/id=id
        self.ip = ip
        self.index_name = index_name
        self.index_type = index_type
        self.es = Elasticsearch([ip])

    def create_index(self):
        if self.es.indices.exists(index=self.index_name):
            self.es.indices.delete(index=self.index_name)
        body = {
            "settings":{
                "index": {
                    "number_of_shards": 1,
                    "number_of_replicas": 1,
                }
            },
            "mappings":{
                "doc":{
                    "properties":{
                        "vus":{
                            "type":"text",
                            "index":"true",
                            "fields": {
                                "keyword": {
                                    "type": "keyword",
                                    "ignore_above": 256
                                }
                            },
                            "index_options": "docs"
                        }
                    }
                }
            }
        }
        self.es.indices.create(index=self.index_name,body=body)

    def insert_data_from_excel(self,file_path):
        tb = xlrd.open_workbook(file_path)
        sheet = tb.sheet_by_index(0)
        rows = sheet.nrows
        for i in range(1, rows):
            print(i)
            app = sheet.cell_value(rowx=i, colx=1)
            string_id = sheet.cell_value(rowx=i, colx=2)
            vus = sheet.cell_value(rowx=i, colx=4)
            vcn = sheet.cell_value(rowx=i, colx=5)
            vtr = sheet.cell_value(rowx=i, colx=6)
            body_insert = {
                "app": app,
                "string_id": string_id,
                "vus": vus,
                "vcn": vcn,
                "vtr": vtr
            }
            self.es.index(index=self.index_name,doc_type=self.index_type,body=body_insert)

    def insert_data_from_sqlite(self):
        conn = sqlite3.connect(db_path)
        cur = conn.cursor()
        cur.execute('select * from string_tbl')
        col_field = [i[0] for i in cur.description]
        count = 0
        while True:
            data = cur.fetchmany(10000)
            print(count)
            if data:
                count += 10000
                bulk_dic = []
                for i in range(len(data)):
                    es_dic = dict(zip(col_field,data[i]))
                    es_dic.update(_index=self.index_name,_type=self.index_type)
                    bulk_dic.append(es_dic)
                helpers.bulk(self.es,bulk_dic)
            else:
                break

    def search_from_excel(self,file_path):
        tb = xlrd.open_workbook(file_path)
        wb = openpyxl.load_workbook(file_path)
        sheet_wb = wb.worksheets[0]
        sheet = tb.sheet_by_index(0)
        rows = sheet.nrows
        # 精确查找
        for i in range(1, rows):
            english = sheet.cell_value(rowx=i, colx=7).strip()
            E = sheet.cell_value(rowx=i, colx=4)
            if E:
                continue
            conn = sqlite3.connect(db_path)
            cur = conn.cursor()
            cur.execute("select * from string_tbl where vus='%s';" % english.replace('\'','\'\''))
            for row in cur:
                sheet_wb[f'E{i+1}'] = row[2]
                sheet_wb[f'F{i+1}'] = row[4]
                sheet_wb[f'G{i+1}'] = row[3]
                wb.save(file_path)

        # 模糊查找
        tb = xlrd.open_workbook(file_path)
        wb = openpyxl.load_workbook(file_path)
        sheet_wb = wb.worksheets[0]
        sheet = tb.sheet_by_index(0)
        rows = sheet.nrows
        for i in range(1, rows):
            english = re.sub(r'\\|\'|\"|\.|\?|:|-|–|!|,|\(|\)|<|>|/',' ',sheet.cell_value(rowx=i, colx=7).strip())
            E = sheet.cell_value(rowx=i, colx=4)
            if E:
                continue
            body_search = {
                "query": {
                    "query_string": {
                        "default_field": "vus",
                        "query": english
                    }
                },
                "size": 1
            }
            try:
                res = self.es.search(index=self.index_name, doc_type=self.index_type, body=body_search)
                for hit in res['hits']['hits']:
                    sheet_wb[f'E{i+1}'] = hit['_source']['string_id']
                    sheet_wb[f'F{i+1}'] = hit['_source']['vcn']
                    sheet_wb[f'G{i+1}'] = hit['_source']['vus']
                    fill = PatternFill("solid", fgColor="FF0000")
                    sheet_wb[f'G{i+1}'].fill = fill
                    wb.save(file_path)
            except RequestError as e:
                print(e)


if __name__ == '__main__':
    obj = ElasticObj('127.0.0.1:9200','python_es01','doc')
    # obj.create_index()
    # ## obj.insert_data_from_excel(r'D:\project\vivo自动工具\sqlite3\客户端字符串.xls')
    # obj.insert_data_from_sqlite()
    # 提交系统的报告
    obj.search_from_excel(final_path)
    print('=' * 25 + 'finish' + '=' * 25)
#!/usr/bin/env python
# -*- coding:utf-8 -*-
# *************************************
# @Time    : 2020/4/11 14:45
# @Author  :
# @Desc    :
# @File    : hyperlinkTool.py
# *************************************
import openpyxl,os,sqlite3,zipfile
from openpyxl.styles import PatternFill


def pre_sql():
    conn = sqlite3.connect('language.db')
    cur = conn.cursor()
    createtb_sql = """create table tb(
    id integer primary key autoincrement,
    language text,
    code text,
    zh text);"""
    cur.execute(createtb_sql)
    wb = openpyxl.load_workbook(r'D:\project\vivo自动工具\cfg\对照表.xlsx')
    sheet = wb['Sheet1']
    rows = sheet.max_row
    cols = sheet.max_column
    for i in range(1,rows+1):
        code = sheet.cell(row=i, column=5).value
        if code:
            zh = sheet.cell(row=i, column=2).value
            language = sheet.cell(row=i, column=3).value
            insert_sql = f'insert into tb(language,code,zh) values("{language}","{code}","{zh}")'
            cur.execute(insert_sql)
    conn.commit()
    cur.close()
    conn.close()


def query_language_by_code(code):
    conn = sqlite3.connect(os.path.join(os.getcwd(),'language.db'))
    cur = conn.cursor()
    cur.execute('select language from tb where code = "%s"' % code)
    try:
        language = cur.fetchone()[0]
        cur.close()
        conn.close()
        return language
    except TypeError:
        print(f'语言代号{code}未找到，请更新sqlite数据！')
        os.system('pause')


def xls_operation(file_path):
    wb = openpyxl.load_workbook(file_path)
    # Translation error表第1行第st_idx列数据
    sheet = wb['Translation error']
    rows = sheet.max_row
    cols = sheet.max_column
    st_idx = None
    l_idx = None
    fill = PatternFill("solid", fgColor="FFFF00")
    for col in range(1,cols+1):
        if '截图\nScreenshot' == sheet.cell(row=1, column=col).value:
            st_idx = col
        if '语言\nLanguage' == sheet.cell(row=1, column=col).value:
            l_idx = col
    if not st_idx or not l_idx:
        print('读取表格数据失败，错误的模板！')
        os.system('pause')
    else:
        for i in range(2,rows+1):
            img = sheet.cell(row=i, column=st_idx).value
            if img:
                code = sheet.cell(row=i, column=l_idx).value
                language = query_language_by_code(code)
                splice_dir = os.path.join(os.getcwd(), language, img[1:-4])
                if os.path.isdir(splice_dir):
                    link = splice_dir
                else:
                    print(f'路径{splice_dir}不存在！！')
                    link = None
            else:
                continue
            if link:
                sheet.cell(row=i, column=st_idx).hyperlink = link
            else:
                sheet.cell(row=i, column=st_idx).fill = fill
    wb.save(file_path)


if __name__ == '__main__':

    root_dir = os.getcwd()
    files = os.listdir(root_dir)
    if 'language.db' not in files:
        print('数据库文件丢失！！')
        os.system('pause')
    xls_files = list(filter(lambda file: file.endswith('.xlsx'), files))
    if not xls_files:
        print('当前执行目录下没有.xlsx文件！')
        os.system('pause')
    else:
        for xls_file in xls_files:
            print('\n'+'-'*15+xls_file)
            try:
                xls_operation(xls_file)
            except zipfile.BadZipFile:
                print('zipfile.BadZipFile: File is not a zip file')
    print('='*15+'转换完成')
    os.system('pause')